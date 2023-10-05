import os
import openpyxl as opx
from colorama import Fore, Style, Back, init

from Entity.Browser import Browser
from func.alterSystem import *

init()

filePrefeitura = {}
nomeUnidade = ""

print("ATTCATBOT, v2.0\n")

while True:
    while True:
        try:
            print("\n\n\n" + Back.WHITE + Fore.BLACK + "------- Arquivo prefeitura -------" + Style.RESET_ALL)
            filePrefeitura = {
                "fName": input("Digite o nome do arquivo: "),
                "ixTabela": int(input("Digite o índice da tabela: ")),
                "colunaCodCliente": int(input("Digite o índice da coluna que contenha o código do cliente: ")),
                "colunaNomeCliente": int(input("Digite o índice da coluna que contenha o nome do cliente: ")),
                "colunaUnidade": int(input("Digite o índice da coluna que contenha a unidade do item: ")),
                "colunaQuantidade": int(input("Digite o índice da coluna que contenha a quantidade: "))
            }
            break
        
        except:
            print("Digite somente dados válidos...")

    if (os.path.isfile("data/{}".format(filePrefeitura["fName"]))):
        try:
            catologo = opx.load_workbook(filename="data/{}".format(filePrefeitura["fName"]))
            tabelaEscolha = catologo[catologo.sheetnames[filePrefeitura["ixTabela"]]]
            nomeUnidade = tabelaEscolha.cell(row=1, column=filePrefeitura["colunaQuantidade"]).value
            print("A unidade que deseja alterar o catálogo é " + Fore.GREEN + str(nomeUnidade).lstrip(" ") + Style.RESET_ALL + "? (y/n) ")
            tacerto = input()
            if tacerto == "y":
                nomeUnidade = str(nomeUnidade).lstrip(" ").replace(" ", "-").lower()
                break
            else:
                print(Fore.RED + 'Por favor, digite novamente os dados...' + Style.RESET_ALL)
        except:
            print(Fore.RED + "Arquivo não suportado pelo sistema..." + Style.RESET_ALL)
    
    else:
        print(Fore.RED + "Arquivo não encontrado, tente novamente..." + Style.RESET_ALL)

navigator = Browser()

navigator.get("https://juizdefora.branetlogistica.com.br/doms/")

navigator.wait_response("Entre no catálogo que deseja fazer a modificações e pressione ENTER")

ChangeCatalogInSystem(navigator)