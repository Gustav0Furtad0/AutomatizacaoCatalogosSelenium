from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from colorama import Fore, Style
import openpyxl as opx

def ChangeCatalogInSystem(browser, path, colQtd, ixtable, colName, colUn, colCod):
    catalogo = opx.load_workbook(filename=path)
    
    def getDataFromArchive(colQtd, ixtable = 0, colName = 2, colUn = 3, colCod = 1):
        ##* Função reutilzada de versões anteriores
        ##! Refatorar código
        
        tabelaEscolha = catalogo[catalogo.sheetnames[ixtable]]
        lista = []
        for i in range(2, tabelaEscolha.max_row):
            line = {
                    "codCliente": tabelaEscolha.cell(row=i, column=colCod).value,
                    "nomeCliente": tabelaEscolha.cell(row=i, column=colName).value,
                    "unidade": tabelaEscolha.cell(row=i, column=colUn).value
                }
            if tabelaEscolha.cell(row=i, column=colQtd).value in [None, "", " ", '\n', '\r', ]:
                line["quantidade"] = 0
            else:
                line["quantidade"] = int(tabelaEscolha.cell(row=i, column=colQtd).value)
            
            
            if tabelaEscolha.cell(row=i, column=colCod).value != None:
                lista.append(line)
            
        return tabelaEscolha
    
    tabelaEscolha = getDataFromArchive(colQtd, ixtable, colName, colUn, colCod)
    
    select = Select(browser.find_element(By.XPATH, "/html/body/div[@id='div_principal']/div[@id='div_principal']/div[@id='conteudo_template']/span[@id='conteudo']/form[@id='cadastro_itens']/div/div/div/div[1]/select"))
    select.select_by_value('500')
    
    sleep(3)
    
    navigator = browser.find_element(By.XPATH, "/html/body/div/div/div/span/form/div/div/div/div[1]/span[2]")
    navigator = navigator.find_elements(By.TAG_NAME, 'a')
    
    while True:
        ind = None
        
        def analizeItem(codItem):
            for i in range(2, tabelaEscolha.max_row):
                if codItem == tabelaEscolha.cell(row=i, column=colCod).value:
                    return True, i
            return False, i
        
        if input("Alterne a página e pressione ENTER || Pressiona 'n' para sair: ") == 'N':
            break
        
        rows = browser.find_elements(By.XPATH, "/html/body/div[@id='div_principal']/div[@id='div_principal']/div[@id='conteudo_template']/span[@id='conteudo']/form/div/div/div/div/table/tbody/tr")
        
        for row in rows:
            cels = row.find_elements(By.TAG_NAME, 'td')
            if cels[5].text.isdigit():
                campo = cels[1].text.splitlines()
                if not (len(campo) <= 1):
                    itemResponse = {}
                    response, idx = analizeItem(campo[1])
                    ind = idx
                    if response and idx is not None:
                        itemResponse = {
                            "codCliente": tabelaEscolha.cell(row=ind, column=colCod).value,
                            "nomeCliente": tabelaEscolha.cell(row=ind, column=colName).value,
                            "unidade": tabelaEscolha.cell(row=ind, column=colUn).value
                        }
                        if tabelaEscolha.cell(row=ind, column=colQtd).value in [None, "", " ", '\n', '\r', ]:
                            itemResponse["quantidade"] = 0
                        else:
                            itemResponse["quantidade"] = int(tabelaEscolha.cell(row=ind, column=colQtd).value)
                    else:
                        continue
                    
                    ix = 0
                    
                    print("Analisando item: " + cels[2].text)
                    sleep(2)
                    print(Fore.CYAN + "Nome sistema: " + cels[2].text + Style.RESET_ALL + " | " + Fore.GREEN + "Nome Prefeitura: " + itemResponse["nomeCliente"] + Fore.CYAN + "\nQuantidade: " + cels[5].text + Style.RESET_ALL +" | " + Fore.GREEN + "Quantidade e unidade Prefeitura: " + str(itemResponse["quantidade"]) + Style.RESET_ALL)
                    # if not input("Os itens acima são iguais ? (y/n) ") == 'y':
                    #     continue
                    
                    while True:
                        try:
                            btsEdit = cels[13].find_element(By.TAG_NAME, 'div').find_elements(By.TAG_NAME, 'a')
                            btsEdit[0].click()
                            
                            inputQtd = cels[5].find_element(By.TAG_NAME, 'div').find_elements(By.TAG_NAME, 'div')[1].find_element(By.TAG_NAME, 'span').find_elements(By.TAG_NAME, 'input')
                            inputQtd[0].click()
                            inputQtd[0].clear()
                            inputQtd[0].send_keys(itemResponse["quantidade"])

                            inputQtd = cels[8].find_element(By.TAG_NAME, 'div').find_elements(By.TAG_NAME, 'div')[1].find_element(By.TAG_NAME, 'span').find_elements(By.TAG_NAME, 'input')
                            inputQtd[0].click()
                            inputQtd[0].clear()
                            inputQtd[0].send_keys(itemResponse["quantidade"])
                            
                            btsEdit[1].click()
                            
                            tabelaEscolha.delete_rows(ind)
                            
                            break
                        
                        except:
                            if ix >= 5:
                                break
                            ix += 1
                            sleep(1)
        
        navigator = browser.find_element(By.XPATH, "/html/body/div/div/div/span/form/div/div/div/div[1]/span[2]")
        navigator = navigator.find_elements(By.TAG_NAME, 'a')
        
    novo_path = "result.xlsx"
    catalogo.save(novo_path)
    catalogo.close()
    